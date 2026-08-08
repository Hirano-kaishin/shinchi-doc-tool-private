from __future__ import annotations

import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook

from agents import AgentService
from config import Config
from storage import ProjectStore


app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": Config.FRONTEND_ORIGIN}})

store = ProjectStore(Config.DATA_DIR)
agents = AgentService(
    api_key=Config.ANTHROPIC_API_KEY,
    model=Config.CLAUDE_MODEL,
    enabled=Config.ENABLE_REMOTE_LLM,
)

SEKOU_SECTION_TITLES = [
    "表紙",
    "工事概要",
    "工事内容",
    "施工場所",
    "計画工程表",
    "現場組織票",
    "施工体系図",
    "使用機械",
    "施工方法",
    "防災対策",
    "防災対策②",
    "防災対策中止基準",
    "施工管理計画",
    "交通管理",
    "現場環境対策",
    "中止基準",
]

ORDERER_SPECS_DIR = Config.DATA_DIR / "orderer_specs"
ORDERER_SPECS_INDEX = ORDERER_SPECS_DIR / "index.json"


def _safe_segment(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]", "_", str(text or "").strip())
    return cleaned[:80] or "unknown"


def _safe_filename(filename: str) -> str:
    name = Path(filename or "file").name
    stem = _safe_segment(Path(name).stem)[:80]
    suffix = "".join(ch for ch in Path(name).suffix if ch.isalnum() or ch == ".")[:10]
    return f"{stem}{suffix}" if stem else f"file{suffix}"


def _load_orderer_specs_index() -> dict:
    ORDERER_SPECS_DIR.mkdir(parents=True, exist_ok=True)
    if not ORDERER_SPECS_INDEX.exists():
        ORDERER_SPECS_INDEX.write_text("{}", encoding="utf-8")
    try:
        return json.loads(ORDERER_SPECS_INDEX.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def _save_orderer_specs_index(data: dict) -> None:
    ORDERER_SPECS_DIR.mkdir(parents=True, exist_ok=True)
    ORDERER_SPECS_INDEX.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _mock_extract() -> dict:
    return {
        "id": uuid.uuid4().hex[:12],
        "upload_id": uuid.uuid4().hex[:12],
        "koji_gaiyo": {
            "koji_mei": "サンプル工事（AI抽出モック）",
            "koji_bango": "SAMPLE-001",
            "hacchusha": "民間発注者",
            "kasho": "千葉県袖ケ浦市",
            "kouki": "2026/04-2026/09",
        },
        "koshu_list": [
            {"koshu": "土工", "saimoku": "掘削", "suryo": 3326, "tani": "m3"},
            {"koshu": "排水路工", "saimoku": "側溝", "suryo": 120, "tani": "m"},
            {"koshu": "発電所設備", "saimoku": "架台・パネル", "suryo": 9, "tani": "区画"},
        ],
        "created_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }


@app.get("/api/health")
def health():
    return jsonify({"ok": True, "service": "shinchi-doc-tool-python"})


@app.get("/api/config")
def config_info():
    return jsonify(
        {
            "claude_model": Config.CLAUDE_MODEL,
            "remote_llm_enabled": Config.ENABLE_REMOTE_LLM,
            "data_dir": str(Path(Config.DATA_DIR).resolve()),
        }
    )


@app.get("/api/projects")
def list_projects():
    return jsonify({"items": store.list()})


@app.get("/api/projects/<pid>")
def get_project(pid: str):
    item = store.get(pid)
    if not item:
        return jsonify({"error": "not_found"}), 404
    return jsonify(item)


@app.post("/api/projects")
def save_project():
    payload = request.get_json(silent=True) or {}
    pid = payload.get("id")
    saved = store.upsert(payload, pid=pid)
    return jsonify(saved)


@app.delete("/api/projects/<pid>")
def delete_project(pid: str):
    ok = store.delete(pid)
    if not ok:
        return jsonify({"error": "not_found"}), 404
    return jsonify({"ok": True})


@app.post("/api/extract")
def extract():
    # This endpoint is intentionally compatible with future multipart uploads.
    # For now, it returns deterministic mock extraction data.
    _ = request.files.getlist("pdf")
    return jsonify(_mock_extract())


@app.post("/api/agent/run")
def run_agent():
    body = request.get_json(silent=True) or {}
    task = body.get("task")
    payload = body.get("payload", {})
    if not task:
        return jsonify({"error": "task_required"}), 400
    try:
        result = agents.run(task, payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify({"task": task, "result": result})


@app.post("/api/orderers/<orderer_id>/spec-files")
def upload_orderer_spec_files(orderer_id: str):
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "files_required"}), 400

    safe_orderer_id = _safe_segment(orderer_id)
    orderer_dir = ORDERER_SPECS_DIR / safe_orderer_id
    orderer_dir.mkdir(parents=True, exist_ok=True)

    index = _load_orderer_specs_index()
    rows = index.get(safe_orderer_id, [])
    uploaded_rows = []

    for f in files:
        if not f or not f.filename:
            continue
        file_id = uuid.uuid4().hex[:12]
        original_name = Path(f.filename).name
        stored_name = f"{file_id}_{_safe_filename(original_name)}"
        target = orderer_dir / stored_name
        f.save(target)
        size = target.stat().st_size
        row = {
            "id": file_id,
            "originalName": original_name,
            "storedName": stored_name,
            "relativePath": str(Path("orderer_specs") / safe_orderer_id / stored_name),
            "size": size,
            "uploadedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        }
        rows.append(row)
        uploaded_rows.append(row)

    index[safe_orderer_id] = rows
    _save_orderer_specs_index(index)
    return jsonify({"orderer_id": safe_orderer_id, "items": uploaded_rows})


@app.get("/api/orderers/<orderer_id>/spec-files")
def list_orderer_spec_files(orderer_id: str):
    safe_orderer_id = _safe_segment(orderer_id)
    index = _load_orderer_specs_index()
    return jsonify({"orderer_id": safe_orderer_id, "items": index.get(safe_orderer_id, [])})


@app.delete("/api/orderers/<orderer_id>/spec-files/<file_id>")
def delete_orderer_spec_file(orderer_id: str, file_id: str):
    safe_orderer_id = _safe_segment(orderer_id)
    safe_file_id = _safe_segment(file_id)
    index = _load_orderer_specs_index()
    rows = index.get(safe_orderer_id, [])
    hit = next((r for r in rows if r.get("id") == safe_file_id), None)
    if not hit:
        return jsonify({"error": "not_found"}), 404

    target = ORDERER_SPECS_DIR / safe_orderer_id / hit.get("storedName", "")
    if target.exists() and target.is_file():
        target.unlink()

    rows = [r for r in rows if r.get("id") != safe_file_id]
    index[safe_orderer_id] = rows
    _save_orderer_specs_index(index)
    return jsonify({"ok": True, "orderer_id": safe_orderer_id, "file_id": safe_file_id})


@app.post("/api/keikakusho/genba-joken")
def genba_joken():
    payload = request.get_json(silent=True) or {}
    result = agents.run("ground_conditions", payload)
    return jsonify(result)


@app.post("/api/keikakusho/schedule")
def keikaku_schedule():
    payload = request.get_json(silent=True) or {}
    result = agents.run("schedule", payload)
    return jsonify(result)


@app.post("/api/keikakusho/sections")
def keikaku_sections():
    payload = request.get_json(silent=True) or {}
    section = payload.get("section")
    project_type = payload.get("project_type", "private_solar")
    context = payload.get("context", {})
    if section:
        result = agents.run(
            "section_draft",
            {
                "section": section,
                "project_type": project_type,
                "context": context,
            },
        )
    else:
        result = agents.run(
            "section_bundle",
            {
                "project_type": project_type,
                "context": context,
            },
        )
    return jsonify(result)


@app.get("/api/keikakusho/section-profile")
def keikaku_section_profile():
    project_type = request.args.get("project_type", "private_solar")
    result = agents.run("section_bundle", {"project_type": project_type, "context": {}})
    return jsonify(
        {
            "project_type": result.get("project_type"),
            "profile": result.get("profile", {}),
        }
    )


@app.post("/api/export/zip")
def export_zip():
    payload = request.get_json(silent=True) or {}
    project = payload.get("project", {})
    schedules = payload.get("schedules", {})
    flow = payload.get("flow", {})
    company_master = payload.get("companyMaster", {})
    orderer_masters = payload.get("ordererMasters", [])
    project_orderer_map = payload.get("projectOrdererMap", {})

    project_id = project.get("id", "project")
    project_name = project.get("name", "工事")
    tasks = schedules.get(project_id, [])
    confirmed_sections = (
        flow.get("sekou", {}).get(project_id, {}).get("confirmedSections", {})
    )

    def _find_selected_orderer() -> dict:
        selected_id = project_orderer_map.get(project_id, "")
        if not selected_id:
            return {}
        for item in orderer_masters:
            if isinstance(item, dict) and item.get("id") == selected_id:
                return item
        return {}

    orderer = _find_selected_orderer()
    orderer_type_map = {
        "private": "民間企業",
        "public_miyazaki": "公共（宮崎県）",
        "public_miyakonojo": "公共（都城市）",
        "public_other": "公共（その他）",
    }

    def _make_meta_wb() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "案件情報"
        ws.append(["項目", "値"])
        ws.append(["工事ID", project_id])
        ws.append(["工事名", project_name])
        ws.append(["施工場所", project.get("place", "")])
        ws.append(["工期", project.get("period", "")])
        ws.append(["受注者", company_master.get("name", "")])
        ws.append(["代表者", company_master.get("representative", "")])
        ws.append(["発注者", orderer.get("name", "")])
        ws.append([
            "発注者区分",
            orderer_type_map.get(orderer.get("type", ""), ""),
        ])
        ws.append(["提出先部署", orderer.get("dept", "")])
        ws.append(["案件コード", orderer.get("code", "")])
        ws.append(["仕様書/管理基準ファイル", orderer.get("specFiles", "")])
        ws.append(["発注者条件", orderer.get("manualRules", "")])
        ws.append(["企業標準文面", company_master.get("standardText", "")])

        ws.append(["施工計画書 決定数", len(confirmed_sections.keys())])

        b = BytesIO()
        wb.save(b)
        return b.getvalue()

    def _make_schedule_wb() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "工程管理表"
        ws.append(["工事ID", project_id])
        ws.append(["工事名", project_name])
        ws.append(["受注者", company_master.get("name", "")])
        ws.append(["発注者", orderer.get("name", "")])
        ws.append(["発注者区分", orderer_type_map.get(orderer.get("type", ""), "")])
        ws.append(["仕様書/管理基準ファイル", orderer.get("specFiles", "")])
        ws.append(["発注者条件", orderer.get("manualRules", "")])
        ws.append([])
        ws.append(["工種", "単位", "数量", "進捗(%)"])
        for row in tasks:
            ws.append([
                row[0] if len(row) > 0 else "",
                row[1] if len(row) > 1 else "",
                row[2] if len(row) > 2 else "",
                row[4] if len(row) > 4 else 0,
            ])

        b = BytesIO()
        wb.save(b)
        return b.getvalue()

    def _make_section_wb(index: int, section_title: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "施工計画書"
        ws.append(["項目", "値"])
        ws.append(["章番号", f"{index:02d}"])
        ws.append(["章名", section_title])
        ws.append(["工事ID", project_id])
        ws.append(["工事名", project_name])
        ws.append(["施工場所", project.get("place", "")])
        ws.append(["工期", project.get("period", "")])
        ws.append(["受注者", company_master.get("name", "")])
        ws.append(["代表者", company_master.get("representative", "")])
        ws.append(["発注者", orderer.get("name", "")])
        ws.append([
            "発注者区分",
            orderer_type_map.get(orderer.get("type", ""), ""),
        ])
        ws.append(["提出先部署", orderer.get("dept", "")])
        ws.append(["案件コード", orderer.get("code", "")])
        ws.append(["仕様書/管理基準ファイル", orderer.get("specFiles", "")])
        ws.append(["発注者条件", orderer.get("manualRules", "")])
        ws.append(["企業標準文面", company_master.get("standardText", "")])
        status = "決定済み" if confirmed_sections.get(section_title) else "未決定"
        ws.append(["状態", status])

        if section_title == "計画工程表":
            ws.append([])
            ws.append(["工種", "単位", "数量", "進捗(%)"])
            for row in tasks:
                ws.append([
                    row[0] if len(row) > 0 else "",
                    row[1] if len(row) > 1 else "",
                    row[2] if len(row) > 2 else "",
                    row[4] if len(row) > 4 else 0,
                ])

        b = BytesIO()
        wb.save(b)
        return b.getvalue()

    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w", compression=ZIP_DEFLATED) as zf:
        zf.writestr("00_案件情報.xlsx", _make_meta_wb())
        zf.writestr("01_工程管理表.xlsx", _make_schedule_wb())
        for idx, section_title in enumerate(SEKOU_SECTION_TITLES, start=1):
            section_name = f"sekou/{idx:02d}_{section_title}.xlsx"
            zf.writestr(section_name, _make_section_wb(idx, section_title))

    zip_buffer.seek(0)
    filename = f"{project_id}_documents.zip"
    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(host=Config.HOST, port=Config.PORT, debug=Config.DEBUG)
